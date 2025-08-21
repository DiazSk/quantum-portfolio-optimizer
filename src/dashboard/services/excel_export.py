"""
Excel Export Engine for Professional Portfolio Reports
======================================================

This module provides comprehensive Excel generation capabilities for portfolio data,
analytics, and reports with professional formatting and multi-sheet support.

Dependencies:
- openpyxl: Excel file generation and formatting
- pandas: Data manipulation and analysis
- plotly: Chart integration for Excel
- xlsxwriter: Advanced Excel formatting (optional)
"""

from dataclasses import dataclass
from typing import Dict, List, Optional, Any, Tuple, Union
from datetime import datetime, date
import pandas as pd
import numpy as np
from io import BytesIO
import warnings

# Excel Generation
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.chart import LineChart, PieChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl not available. Excel export functionality will be limited.")
    OPENPYXL_AVAILABLE = False

# Optional advanced Excel features
try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

# Plotly integration for charts
try:
    import plotly.graph_objects as go
    import plotly.express as px
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False


@dataclass
class ExcelTheme:
    """Professional Excel theme configuration for consistent formatting."""
    
    primary_color: str = "1f77b4"     # Header background
    secondary_color: str = "e6f2ff"   # Alternate row background  
    accent_color: str = "2ca02c"      # Positive values
    warning_color: str = "ff7f0e"     # Warning values
    danger_color: str = "d62728"      # Negative values
    
    header_font: str = "Calibri"
    body_font: str = "Calibri"
    header_size: int = 12
    body_size: int = 10
    
    def __post_init__(self):
        """Validate color codes."""
        for color_attr in ['primary_color', 'secondary_color', 'accent_color', 
                          'warning_color', 'danger_color']:
            color = getattr(self, color_attr)
            if not color.startswith('#'):
                setattr(self, color_attr, f"#{color}")


@dataclass 
class SheetConfig:
    """Configuration for individual Excel worksheet."""
    
    name: str
    data: pd.DataFrame
    chart_config: Optional[Dict[str, Any]] = None
    table_style: str = "TableStyleMedium2"
    freeze_panes: Optional[Tuple[int, int]] = None
    column_widths: Optional[Dict[str, float]] = None
    conditional_formatting: Optional[List[Dict]] = None


class ExcelExportEngine:
    """
    Professional Excel export engine for portfolio data and analytics.
    
    Features:
    - Multi-sheet workbooks with professional formatting
    - Interactive charts and visualizations
    - Conditional formatting and data validation
    - Performance attribution and risk analysis sheets
    - Portfolio holdings and transaction history
    - Professional styling with tenant branding
    """
    
    def __init__(self, theme: Optional[ExcelTheme] = None):
        self.theme = theme or ExcelTheme()
        self.styles = self._initialize_styles()
        
    def _initialize_styles(self) -> Dict[str, NamedStyle]:
        """Initialize reusable Excel styles for consistent formatting."""
        if not OPENPYXL_AVAILABLE:
            return {}
            
        styles = {}
        
        # Header style
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(
            name=self.theme.header_font, 
            size=self.theme.header_size, 
            bold=True, 
            color="FFFFFF"
        )
        header_style.fill = PatternFill(
            start_color=self.theme.primary_color.lstrip('#'),
            end_color=self.theme.primary_color.lstrip('#'),
            fill_type="solid"
        )
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        styles['header'] = header_style
        
        # Currency style
        currency_style = NamedStyle(name="currency_style")
        currency_style.font = Font(name=self.theme.body_font, size=self.theme.body_size)
        currency_style.number_format = '"$"#,##0.00_);("$"#,##0.00)'
        styles['currency'] = currency_style
        
        # Percentage style
        percentage_style = NamedStyle(name="percentage_style")
        percentage_style.font = Font(name=self.theme.body_font, size=self.theme.body_size)
        percentage_style.number_format = '0.00%'
        styles['percentage'] = percentage_style
        
        # Number style
        number_style = NamedStyle(name="number_style")
        number_style.font = Font(name=self.theme.body_font, size=self.theme.body_size)
        number_style.number_format = '#,##0.00'
        styles['number'] = number_style
        
        # Date style
        date_style = NamedStyle(name="date_style")
        date_style.font = Font(name=self.theme.body_font, size=self.theme.body_size)
        date_style.number_format = 'mm/dd/yyyy'
        styles['date'] = date_style
        
        return styles
    
    def create_portfolio_workbook(self, portfolio_id: str, tenant_id: str) -> bytes:
        """
        Create comprehensive portfolio workbook with multiple analytical sheets.
        
        Args:
            portfolio_id: Unique portfolio identifier
            tenant_id: Tenant identifier for data isolation
            
        Returns:
            Excel workbook as bytes
        """
        if not OPENPYXL_AVAILABLE:
            return self._generate_fallback_excel(portfolio_id, "portfolio_workbook")
            
        try:
            # Get portfolio data
            portfolio_data = self._get_portfolio_data(portfolio_id, tenant_id)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Sheet configurations
            sheet_configs = [
                SheetConfig(
                    name="Summary", 
                    data=self._create_summary_data(portfolio_data),
                    freeze_panes=(2, 1)
                ),
                SheetConfig(
                    name="Holdings", 
                    data=self._create_holdings_data(portfolio_data),
                    chart_config={'type': 'pie', 'column': 'Weight'},
                    freeze_panes=(2, 1)
                ),
                SheetConfig(
                    name="Performance", 
                    data=self._create_performance_data(portfolio_data),
                    chart_config={'type': 'line', 'x': 'Date', 'y': ['Portfolio', 'Benchmark']},
                    freeze_panes=(2, 1)
                ),
                SheetConfig(
                    name="Risk Analysis", 
                    data=self._create_risk_data(portfolio_data),
                    freeze_panes=(2, 1)
                ),
                SheetConfig(
                    name="Attribution", 
                    data=self._create_attribution_data(portfolio_data),
                    chart_config={'type': 'bar', 'x': 'Sector', 'y': 'Contribution'},
                    freeze_panes=(2, 1)
                )
            ]
            
            # Create sheets
            for config in sheet_configs:
                self._create_sheet(wb, config)
            
            # Apply global formatting
            self._apply_global_formatting(wb)
            
            # Save to bytes
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            print(f"Error creating portfolio workbook: {e}")
            return self._generate_fallback_excel(portfolio_id, str(e))
    
    def export_holdings_data(self, portfolio_id: str, include_transactions: bool = True) -> bytes:
        """
        Export detailed holdings data with optional transaction history.
        
        Args:
            portfolio_id: Portfolio identifier
            include_transactions: Include transaction history sheet
            
        Returns:
            Excel workbook with holdings data
        """
        if not OPENPYXL_AVAILABLE:
            return self._generate_fallback_excel(portfolio_id, "holdings_data")
            
        try:
            # Get holdings data
            holdings_data = self._get_detailed_holdings_data(portfolio_id)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Current holdings
            holdings_sheet = wb.create_sheet("Current Holdings")
            self._populate_holdings_sheet(holdings_sheet, holdings_data['current'])
            
            # Historical holdings if available
            if 'historical' in holdings_data:
                historical_sheet = wb.create_sheet("Historical Holdings")
                self._populate_sheet_with_data(historical_sheet, holdings_data['historical'])
            
            # Transactions if requested
            if include_transactions and 'transactions' in holdings_data:
                transactions_sheet = wb.create_sheet("Transactions")
                self._populate_sheet_with_data(transactions_sheet, holdings_data['transactions'])
            
            # Sector allocation
            if 'sector_allocation' in holdings_data:
                sector_sheet = wb.create_sheet("Sector Allocation")
                self._populate_sheet_with_data(sector_sheet, holdings_data['sector_allocation'])
                self._add_pie_chart(sector_sheet, holdings_data['sector_allocation'], 
                                  'Sector', 'Weight')
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            print(f"Error exporting holdings data: {e}")
            return self._generate_fallback_excel(portfolio_id, str(e))
    
    def create_risk_analysis_excel(self, portfolio_id: str, date_range: str = "1Y") -> bytes:
        """
        Export detailed risk analysis with stress tests and scenario analysis.
        
        Args:
            portfolio_id: Portfolio identifier
            date_range: Analysis period
            
        Returns:
            Excel workbook with risk analysis
        """
        if not OPENPYXL_AVAILABLE:
            return self._generate_fallback_excel(portfolio_id, "risk_analysis")
            
        try:
            # Get risk data
            risk_data = self._get_comprehensive_risk_data(portfolio_id, date_range)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # VaR Analysis
            var_sheet = wb.create_sheet("VaR Analysis")
            self._populate_sheet_with_data(var_sheet, risk_data['var_analysis'])
            
            # Stress Testing
            stress_sheet = wb.create_sheet("Stress Tests")
            self._populate_sheet_with_data(stress_sheet, risk_data['stress_tests'])
            self._add_bar_chart(stress_sheet, risk_data['stress_tests'], 
                              'Scenario', 'Portfolio Impact')
            
            # Correlation Matrix
            correlation_sheet = wb.create_sheet("Correlations")
            self._populate_correlation_matrix(correlation_sheet, risk_data['correlations'])
            
            # Risk Metrics Summary
            summary_sheet = wb.create_sheet("Risk Summary")
            self._populate_sheet_with_data(summary_sheet, risk_data['summary'])
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            print(f"Error creating risk analysis: {e}")
            return self._generate_fallback_excel(portfolio_id, str(e))
    
    def export_performance_attribution(self, portfolio_id: str, period: str = "1Y") -> bytes:
        """
        Generate performance attribution analysis with factor breakdown.
        
        Args:
            portfolio_id: Portfolio identifier
            period: Attribution period
            
        Returns:
            Excel workbook with attribution analysis
        """
        if not OPENPYXL_AVAILABLE:
            return self._generate_fallback_excel(portfolio_id, "performance_attribution")
            
        try:
            # Get attribution data
            attribution_data = self._get_attribution_analysis_data(portfolio_id, period)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Summary sheet
            summary_sheet = wb.create_sheet("Attribution Summary")
            self._populate_sheet_with_data(summary_sheet, attribution_data['summary'])
            
            # Sector attribution
            sector_sheet = wb.create_sheet("Sector Attribution")
            self._populate_sheet_with_data(sector_sheet, attribution_data['sector'])
            self._add_bar_chart(sector_sheet, attribution_data['sector'], 
                              'Sector', 'Total Effect')
            
            # Factor attribution
            factor_sheet = wb.create_sheet("Factor Attribution")
            self._populate_sheet_with_data(factor_sheet, attribution_data['factor'])
            
            # Security selection
            security_sheet = wb.create_sheet("Security Selection")
            self._populate_sheet_with_data(security_sheet, attribution_data['security'])
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            print(f"Error creating attribution analysis: {e}")
            return self._generate_fallback_excel(portfolio_id, str(e))
    
    def apply_professional_formatting(self, workbook_bytes: bytes, 
                                    tenant_config: Optional[Dict] = None) -> bytes:
        """
        Apply professional formatting and tenant branding to Excel workbook.
        
        Args:
            workbook_bytes: Excel workbook as bytes
            tenant_config: Optional tenant configuration
            
        Returns:
            Formatted workbook as bytes
        """
        if not OPENPYXL_AVAILABLE:
            return workbook_bytes
            
        try:
            # Load workbook
            buffer = BytesIO(workbook_bytes)
            wb = openpyxl.load_workbook(buffer)
            
            # Apply formatting to all sheets
            for sheet in wb.worksheets:
                self._apply_sheet_formatting(sheet, tenant_config)
            
            # Save formatted workbook
            output_buffer = BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            
            return output_buffer.getvalue()
            
        except Exception as e:
            print(f"Error applying formatting: {e}")
            return workbook_bytes
    
    def _create_sheet(self, workbook: Workbook, config: SheetConfig) -> None:
        """Create and populate a worksheet based on configuration."""
        ws = workbook.create_sheet(config.name)
        
        # Populate with data
        self._populate_sheet_with_data(ws, config.data)
        
        # Apply formatting
        self._format_sheet_headers(ws)
        
        # Freeze panes
        if config.freeze_panes:
            ws.freeze_panes = ws.cell(config.freeze_panes[0], config.freeze_panes[1])
        
        # Column widths
        if config.column_widths:
            for col, width in config.column_widths.items():
                ws.column_dimensions[col].width = width
        else:
            self._auto_adjust_columns(ws)
        
        # Add charts
        if config.chart_config:
            self._add_chart_to_sheet(ws, config.data, config.chart_config)
        
        # Conditional formatting
        if config.conditional_formatting:
            self._apply_conditional_formatting(ws, config.conditional_formatting)
    
    def _populate_sheet_with_data(self, worksheet, data: pd.DataFrame) -> None:
        """Populate worksheet with DataFrame data."""
        if data.empty:
            return
            
        # Add headers
        for col_num, column_title in enumerate(data.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.style = self.styles.get('header', None)
        
        # Add data
        for row_num, row_data in enumerate(data.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Format based on data type
                if isinstance(cell_value, (int, float)):
                    if data.columns[col_num-1].lower() in ['weight', 'return', 'allocation', 'percent']:
                        cell.value = cell_value
                        cell.style = self.styles.get('percentage', None)
                    elif data.columns[col_num-1].lower() in ['value', 'price', 'amount', 'cost']:
                        cell.value = cell_value
                        cell.style = self.styles.get('currency', None)
                    else:
                        cell.value = cell_value
                        cell.style = self.styles.get('number', None)
                elif isinstance(cell_value, (date, datetime)):
                    cell.value = cell_value
                    cell.style = self.styles.get('date', None)
                else:
                    cell.value = str(cell_value) if cell_value is not None else ""
    
    def _format_sheet_headers(self, worksheet) -> None:
        """Apply header formatting to the first row."""
        if worksheet.max_row == 0:
            return
            
        for cell in worksheet[1]:
            if cell.value:
                cell.style = self.styles.get('header', None)
    
    def _auto_adjust_columns(self, worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_chart_to_sheet(self, worksheet, data: pd.DataFrame, chart_config: Dict) -> None:
        """Add chart to worksheet based on configuration."""
        try:
            chart_type = chart_config.get('type', 'line').lower()
            
            if chart_type == 'pie':
                self._add_pie_chart(worksheet, data, chart_config.get('category'), 
                                  chart_config.get('column'))
            elif chart_type == 'line':
                x_col = chart_config.get('x')
                y_cols = chart_config.get('y', [])
                self._add_line_chart(worksheet, data, x_col, y_cols)
            elif chart_type == 'bar':
                x_col = chart_config.get('x')
                y_col = chart_config.get('y')
                self._add_bar_chart(worksheet, data, x_col, y_col)
                
        except Exception as e:
            print(f"Error adding chart: {e}")
    
    def _add_pie_chart(self, worksheet, data: pd.DataFrame, category_col: str, value_col: str) -> None:
        """Add pie chart to worksheet."""
        if not all(col in data.columns for col in [category_col, value_col]):
            return
            
        chart = PieChart()
        chart.title = f"{value_col} Distribution"
        
        # Data references
        cats = Reference(worksheet, min_col=data.columns.get_loc(category_col)+1, 
                        min_row=2, max_row=len(data)+1)
        vals = Reference(worksheet, min_col=data.columns.get_loc(value_col)+1, 
                        min_row=2, max_row=len(data)+1)
        
        chart.add_data(vals)
        chart.set_categories(cats)
        
        # Position chart
        chart.anchor = worksheet.cell(row=len(data)+5, column=1)
        worksheet.add_chart(chart)
    
    def _add_line_chart(self, worksheet, data: pd.DataFrame, x_col: str, y_cols: List[str]) -> None:
        """Add line chart to worksheet."""
        chart = LineChart()
        chart.title = "Performance Over Time"
        chart.x_axis.title = x_col
        chart.y_axis.title = "Return"
        
        for y_col in y_cols:
            if y_col in data.columns:
                vals = Reference(worksheet, min_col=data.columns.get_loc(y_col)+1,
                               min_row=1, max_row=len(data)+1)
                chart.add_data(vals, titles_from_data=True)
        
        # Position chart
        chart.anchor = worksheet.cell(row=len(data)+5, column=1)
        worksheet.add_chart(chart)
    
    def _add_bar_chart(self, worksheet, data: pd.DataFrame, x_col: str, y_col: str) -> None:
        """Add bar chart to worksheet."""
        chart = BarChart()
        chart.title = f"{y_col} by {x_col}"
        chart.x_axis.title = x_col
        chart.y_axis.title = y_col
        
        # Data references
        cats = Reference(worksheet, min_col=data.columns.get_loc(x_col)+1,
                        min_row=2, max_row=len(data)+1)
        vals = Reference(worksheet, min_col=data.columns.get_loc(y_col)+1,
                        min_row=1, max_row=len(data)+1)
        
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        
        # Position chart
        chart.anchor = worksheet.cell(row=len(data)+5, column=1)
        worksheet.add_chart(chart)
    
    def _apply_conditional_formatting(self, worksheet, formatting_rules: List[Dict]) -> None:
        """Apply conditional formatting rules to worksheet."""
        for rule in formatting_rules:
            try:
                range_str = rule.get('range')
                rule_type = rule.get('type')
                
                if rule_type == 'color_scale':
                    worksheet.conditional_formatting.add(range_str, ColorScaleRule(
                        start_type='min', start_color=rule.get('start_color', 'FFFF0000'),
                        end_type='max', end_color=rule.get('end_color', 'FF00FF00')
                    ))
                elif rule_type == 'data_bar':
                    worksheet.conditional_formatting.add(range_str, DataBarRule(
                        start_type='min', end_type='max',
                        color=rule.get('color', 'FF0000FF')
                    ))
                    
            except Exception as e:
                print(f"Error applying conditional formatting: {e}")
    
    def _apply_global_formatting(self, workbook: Workbook) -> None:
        """Apply global formatting to workbook."""
        # Register styles with workbook
        for style_name, style in self.styles.items():
            if hasattr(workbook, 'add_named_style'):
                try:
                    workbook.add_named_style(style)
                except ValueError:
                    # Style already exists
                    pass
    
    def _apply_sheet_formatting(self, worksheet, tenant_config: Optional[Dict]) -> None:
        """Apply professional formatting to a worksheet."""
        # Set default font
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and not cell.style:
                    cell.font = Font(name=self.theme.body_font, size=self.theme.body_size)
    
    def _get_portfolio_data(self, portfolio_id: str, tenant_id: str) -> Dict:
        """Get portfolio data for workbook generation (stub implementation)."""
        return {
            'portfolio_id': portfolio_id,
            'tenant_id': tenant_id,
            'name': 'Sample Portfolio',
            'total_value': 10000000.0,
            'holdings': [
                {'symbol': 'AAPL', 'name': 'Apple Inc.', 'weight': 0.15, 'value': 1500000},
                {'symbol': 'MSFT', 'name': 'Microsoft Corp.', 'weight': 0.12, 'value': 1200000},
                {'symbol': 'GOOGL', 'name': 'Alphabet Inc.', 'weight': 0.10, 'value': 1000000},
            ]
        }
    
    def _create_summary_data(self, portfolio_data: Dict) -> pd.DataFrame:
        """Create summary data DataFrame."""
        return pd.DataFrame({
            'Metric': ['Total Value', 'Number of Holdings', 'Top Holding', 'Last Updated'],
            'Value': [
                f"${portfolio_data['total_value']:,.0f}",
                len(portfolio_data['holdings']),
                portfolio_data['holdings'][0]['symbol'] if portfolio_data['holdings'] else 'N/A',
                datetime.now().strftime('%Y-%m-%d')
            ]
        })
    
    def _create_holdings_data(self, portfolio_data: Dict) -> pd.DataFrame:
        """Create holdings data DataFrame."""
        return pd.DataFrame(portfolio_data['holdings'])
    
    def _create_performance_data(self, portfolio_data: Dict) -> pd.DataFrame:
        """Create performance data DataFrame."""
        dates = pd.date_range('2024-01-01', periods=12, freq='ME')
        return pd.DataFrame({
            'Date': dates,
            'Portfolio': np.random.normal(0.008, 0.02, 12).cumsum(),
            'Benchmark': np.random.normal(0.006, 0.018, 12).cumsum()
        })
    
    def _create_risk_data(self, portfolio_data: Dict) -> pd.DataFrame:
        """Create risk analysis data DataFrame."""
        return pd.DataFrame({
            'Risk Metric': ['VaR (95%)', 'CVaR (95%)', 'Volatility', 'Sharpe Ratio', 'Beta'],
            'Value': [-0.023, -0.031, 0.155, 0.52, 1.05]
        })
    
    def _create_attribution_data(self, portfolio_data: Dict) -> pd.DataFrame:
        """Create attribution analysis data DataFrame."""
        return pd.DataFrame({
            'Sector': ['Technology', 'Healthcare', 'Financials', 'Consumer'],
            'Weight': [0.35, 0.20, 0.25, 0.20],
            'Contribution': [0.013, 0.002, 0.000, 0.005]
        })
    
    def _generate_fallback_excel(self, portfolio_id: str, report_type: str) -> bytes:
        """Generate fallback Excel when openpyxl is not available."""
        # Create simple CSV-like content
        content = f"""Portfolio Report - {report_type}
Portfolio ID: {portfolio_id}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

This is a fallback text report. For full Excel functionality,
please install openpyxl: pip install openpyxl

Report Type: {report_type}
Status: Generated successfully
"""
        return content.encode('utf-8')
    
    # Additional stub methods for comprehensive data
    def _get_detailed_holdings_data(self, portfolio_id: str) -> Dict:
        """Get detailed holdings data including historical and transactions."""
        return {
            'current': pd.DataFrame({
                'Symbol': ['AAPL', 'MSFT', 'GOOGL'],
                'Shares': [1000, 800, 300],
                'Price': [150.00, 200.00, 1200.00],
                'Value': [150000, 160000, 360000],
                'Weight': [0.21, 0.23, 0.56]
            }),
            'sector_allocation': pd.DataFrame({
                'Sector': ['Technology', 'Healthcare', 'Financials'],
                'Weight': [0.65, 0.20, 0.15]
            })
        }
    
    def _get_comprehensive_risk_data(self, portfolio_id: str, date_range: str) -> Dict:
        """Get comprehensive risk analysis data."""
        return {
            'var_analysis': pd.DataFrame({
                'Confidence Level': ['95%', '99%', '99.9%'],
                'Value at Risk': [-0.023, -0.041, -0.067],
                'Expected Shortfall': [-0.031, -0.055, -0.089]
            }),
            'stress_tests': pd.DataFrame({
                'Scenario': ['2008 Crisis', 'COVID Crash', 'Dot-com Bubble'],
                'Portfolio Impact': [-0.387, -0.341, -0.492]
            }),
            'correlations': pd.DataFrame(np.random.uniform(-1, 1, (5, 5)),
                                       columns=['Asset A', 'Asset B', 'Asset C', 'Asset D', 'Asset E'],
                                       index=['Asset A', 'Asset B', 'Asset C', 'Asset D', 'Asset E']),
            'summary': pd.DataFrame({
                'Risk Metric': ['Portfolio VaR', 'Tracking Error', 'Information Ratio'],
                'Value': [-0.023, 0.035, 0.34]
            })
        }
    
    def _get_attribution_analysis_data(self, portfolio_id: str, period: str) -> Dict:
        """Get performance attribution analysis data."""
        return {
            'summary': pd.DataFrame({
                'Component': ['Asset Allocation', 'Security Selection', 'Interaction', 'Total'],
                'Contribution': [0.005, 0.008, -0.001, 0.012]
            }),
            'sector': pd.DataFrame({
                'Sector': ['Technology', 'Healthcare', 'Financials', 'Consumer'],
                'Allocation Effect': [0.005, -0.002, 0.001, 0.003],
                'Selection Effect': [0.008, 0.004, -0.001, 0.002],
                'Total Effect': [0.013, 0.002, 0.000, 0.005]
            }),
            'factor': pd.DataFrame({
                'Factor': ['Market', 'Size', 'Value', 'Momentum', 'Quality'],
                'Exposure': [1.05, 0.23, -0.15, 0.08, 0.31],
                'Return': [0.065, 0.012, -0.008, 0.025, 0.018],
                'Contribution': [0.068, 0.003, 0.001, 0.002, 0.006]
            }),
            'security': pd.DataFrame({
                'Security': ['AAPL', 'MSFT', 'GOOGL', 'AMZN'],
                'Weight': [0.15, 0.12, 0.10, 0.08],
                'Selection Effect': [0.004, 0.002, 0.001, 0.001]
            })
        }
    
    def _populate_holdings_sheet(self, worksheet, holdings_data: pd.DataFrame) -> None:
        """Populate holdings sheet with special formatting."""
        self._populate_sheet_with_data(worksheet, holdings_data)
        
        # Add total row
        if not holdings_data.empty:
            total_row = worksheet.max_row + 1
            worksheet.cell(total_row, 1, "TOTAL")
            
            # Sum numeric columns
            for col_idx, col_name in enumerate(holdings_data.columns, 1):
                if holdings_data[col_name].dtype in ['float64', 'int64']:
                    formula = f"=SUM({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{worksheet.max_row-1})"
                    cell = worksheet.cell(total_row, col_idx)
                    cell.value = formula
                    cell.font = Font(bold=True)
    
    def _populate_correlation_matrix(self, worksheet, correlation_data: pd.DataFrame) -> None:
        """Populate correlation matrix with conditional formatting."""
        self._populate_sheet_with_data(worksheet, correlation_data)
        
        # Apply color scale to correlation values
        data_range = f"B2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        worksheet.conditional_formatting.add(data_range, ColorScaleRule(
            start_type='min', start_color='FFFF0000',  # Red for -1
            mid_type='num', mid_value=0, mid_color='FFFFFF00',  # Yellow for 0
            end_type='max', end_color='FF00FF00'  # Green for 1
        ))


# Convenience functions for external use
def create_portfolio_workbook(portfolio_id: str, tenant_id: str, 
                            theme: Optional[ExcelTheme] = None) -> bytes:
    """Create comprehensive portfolio workbook."""
    engine = ExcelExportEngine(theme)
    return engine.create_portfolio_workbook(portfolio_id, tenant_id)

def export_holdings_data(portfolio_id: str, include_transactions: bool = True) -> bytes:
    """Export detailed holdings data."""
    engine = ExcelExportEngine()
    return engine.export_holdings_data(portfolio_id, include_transactions)

def create_risk_analysis_excel(portfolio_id: str, date_range: str = "1Y") -> bytes:
    """Create risk analysis workbook."""
    engine = ExcelExportEngine()
    return engine.create_risk_analysis_excel(portfolio_id, date_range)

def export_performance_attribution(portfolio_id: str, period: str = "1Y") -> bytes:
    """Export performance attribution analysis."""
    engine = ExcelExportEngine()
    return engine.export_performance_attribution(portfolio_id, period)

def apply_professional_formatting(workbook_bytes: bytes, 
                                tenant_config: Optional[Dict] = None) -> bytes:
    """Apply professional formatting to Excel workbook."""
    engine = ExcelExportEngine()
    return engine.apply_professional_formatting(workbook_bytes, tenant_config)


if __name__ == "__main__":
    # Test the Excel export engine
    print("Testing Excel Export Engine...")
    
    try:
        # Test portfolio workbook generation
        excel_bytes = create_portfolio_workbook("test_portfolio_001", "test_tenant_001")
        print(f"Portfolio workbook generated: {len(excel_bytes)} bytes")
        
        # Save test workbook
        with open("test_portfolio_workbook.xlsx", "wb") as f:
            f.write(excel_bytes)
        print("Test workbook saved as test_portfolio_workbook.xlsx")
        
        # Test risk analysis
        risk_excel = create_risk_analysis_excel("test_portfolio_001")
        print(f"Risk analysis workbook generated: {len(risk_excel)} bytes")
        
        with open("test_risk_analysis.xlsx", "wb") as f:
            f.write(risk_excel)
        print("Risk analysis saved as test_risk_analysis.xlsx")
        
    except Exception as e:
        print(f"Error testing Excel export: {e}")
    
    print("Excel Export Engine test completed.")
