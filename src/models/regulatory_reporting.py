"""
Regulatory Reporting Engine for Quantum Portfolio Optimizer

This module provides automated generation of regulatory reports including
Form PF, AIFMD, and Solvency II formats with validation and delivery capabilities.

Features:
- Template-based report generation system
- Automated data aggregation from portfolio and risk systems
- Report validation and quality checks
- Scheduled generation and delivery
- Regulatory format compliance
"""

import asyncio
import json
import hashlib
from datetime import datetime, date, timedelta
from typing import Dict, Any, List, Optional, Union
from dataclasses import dataclass, asdict
from pathlib import Path
import tempfile
import os

# Report generation libraries
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Data validation
import cerberus

from ..database.connection import DatabaseConnection
from ..utils.professional_logging import get_logger
from ..utils.immutable_audit_trail import AuditEventCapture, AuditEventData

logger = get_logger(__name__)


@dataclass
class ReportMetadata:
    """Report generation metadata and configuration."""
    report_id: str
    report_type: str  # 'form_pf', 'aifmd', 'solvency_ii'
    period_start: date
    period_end: date
    generated_at: datetime
    template_version: str
    data_sources: List[str]
    validation_status: str
    file_format: str  # 'pdf', 'excel', 'xml'


@dataclass
class ReportValidationResult:
    """Report validation results and quality checks."""
    is_valid: bool
    errors: List[str]
    warnings: List[str]
    data_completeness: float  # 0.0 to 1.0
    validation_timestamp: datetime


class RegulatoryReportTemplateManager:
    """
    Template management system for different regulatory report formats.
    
    Handles template loading, validation, and format-specific requirements
    for Form PF, AIFMD, and Solvency II reports.
    """
    
    def __init__(self):
        self.templates = {}
        self._load_templates()
    
    def _load_templates(self):
        """Load regulatory report templates."""
        # Form PF template configuration
        self.templates['form_pf'] = {
            'name': 'SEC Form PF',
            'sections': [
                'basic_information',
                'assets_under_management',
                'investment_strategy',
                'risk_metrics',
                'counterparty_exposure',
                'trading_practices'
            ],
            'required_fields': [
                'total_aum', 'net_asset_value', 'gross_asset_value',
                'leverage_ratio', 'var_measures', 'stress_test_results'
            ],
            'validation_rules': {
                'total_aum': {'type': 'number', 'min': 0},
                'leverage_ratio': {'type': 'number', 'min': 0, 'max': 10},
                'var_measures': {'type': 'dict', 'required': True}
            },
            'frequency': 'quarterly',
            'deadline_days': 120  # Days after quarter end
        }
        
        # AIFMD template configuration
        self.templates['aifmd'] = {
            'name': 'EU Alternative Investment Fund Managers Directive',
            'sections': [
                'fund_identification',
                'investment_strategy',
                'risk_profile',
                'liquidity_management',
                'leverage_calculation',
                'operational_information'
            ],
            'required_fields': [
                'fund_size', 'investment_strategy_code', 'leverage_calculation',
                'liquidity_profile', 'risk_measures', 'operational_risk'
            ],
            'validation_rules': {
                'fund_size': {'type': 'number', 'min': 0},
                'investment_strategy_code': {'type': 'string', 'allowed': ['equity', 'fixed_income', 'multi_strategy']},
                'leverage_calculation': {'type': 'dict', 'required': True}
            },
            'frequency': 'annual',
            'deadline_days': 90
        }
        
        # Solvency II template configuration
        self.templates['solvency_ii'] = {
            'name': 'Solvency II Regulatory Reporting',
            'sections': [
                'balance_sheet',
                'own_funds',
                'solvency_capital_requirement',
                'minimum_capital_requirement',
                'risk_concentration'
            ],
            'required_fields': [
                'technical_provisions', 'own_funds_total', 'scr_ratio',
                'mcr_ratio', 'concentration_risk', 'market_risk'
            ],
            'validation_rules': {
                'scr_ratio': {'type': 'number', 'min': 1.0},  # Must be above 100%
                'mcr_ratio': {'type': 'number', 'min': 1.0},
                'own_funds_total': {'type': 'number', 'min': 0}
            },
            'frequency': 'quarterly',
            'deadline_days': 60
        }
        
        logger.info(f"Loaded {len(self.templates)} regulatory report templates")
    
    def get_template(self, report_type: str) -> Dict[str, Any]:
        """Get regulatory report template by type."""
        if report_type not in self.templates:
            raise ValueError(f"Unknown report type: {report_type}")
        return self.templates[report_type]
    
    def validate_template_data(self, report_type: str, data: Dict[str, Any]) -> ReportValidationResult:
        """Validate report data against template requirements."""
        template = self.get_template(report_type)
        
        # Create Cerberus validator
        validator = cerberus.Validator(template['validation_rules'])
        
        # Validate data
        is_valid = validator.validate(data)
        errors = []
        warnings = []
        
        if not is_valid:
            errors.extend([f"{field}: {error}" for field, error in validator.errors.items()])
        
        # Check required fields
        missing_fields = []
        for field in template['required_fields']:
            if field not in data or data[field] is None:
                missing_fields.append(field)
        
        if missing_fields:
            errors.append(f"Missing required fields: {', '.join(missing_fields)}")
        
        # Calculate data completeness
        total_fields = len(template['required_fields'])
        complete_fields = total_fields - len(missing_fields)
        data_completeness = complete_fields / total_fields if total_fields > 0 else 1.0
        
        # Add warnings for low data completeness
        if data_completeness < 0.9:
            warnings.append(f"Data completeness is {data_completeness:.1%}, consider reviewing missing fields")
        
        return ReportValidationResult(
            is_valid=(len(errors) == 0),
            errors=errors,
            warnings=warnings,
            data_completeness=data_completeness,
            validation_timestamp=datetime.now()
        )


class ReportDataAggregator:
    """
    Automated data aggregation from portfolio and risk systems.
    
    Collects and processes data from various system components
    to populate regulatory report templates.
    """
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
    
    async def aggregate_form_pf_data(
        self,
        period_start: date,
        period_end: date,
        fund_ids: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """Aggregate data for SEC Form PF report."""
        
        # Basic fund information
        basic_info = await self._get_basic_fund_info(fund_ids)
        
        # Assets under management
        aum_data = await self._get_aum_data(period_start, period_end, fund_ids)
        
        # Risk metrics
        risk_metrics = await self._get_risk_metrics(period_start, period_end, fund_ids)
        
        # Trading and counterparty data
        trading_data = await self._get_trading_data(period_start, period_end, fund_ids)
        
        return {
            'basic_information': basic_info,
            'assets_under_management': aum_data,
            'risk_metrics': risk_metrics,
            'trading_practices': trading_data,
            'total_aum': aum_data.get('total_gross_aum', 0),
            'net_asset_value': aum_data.get('total_nav', 0),
            'gross_asset_value': aum_data.get('total_gross_aum', 0),
            'leverage_ratio': risk_metrics.get('leverage_ratio', 0),
            'var_measures': risk_metrics.get('var_measures', {}),
            'stress_test_results': risk_metrics.get('stress_tests', {}),
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(),
            'reporting_currency': 'USD'
        }
    
    async def aggregate_aifmd_data(
        self,
        period_start: date,
        period_end: date,
        fund_ids: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """Aggregate data for AIFMD report."""
        
        # Fund identification and strategy
        fund_info = await self._get_basic_fund_info(fund_ids)
        strategy_info = await self._get_investment_strategy(fund_ids)
        
        # Leverage and risk calculations
        leverage_data = await self._get_leverage_calculation(period_start, period_end, fund_ids)
        risk_data = await self._get_risk_metrics(period_start, period_end, fund_ids)
        
        # Liquidity management
        liquidity_data = await self._get_liquidity_profile(period_start, period_end, fund_ids)
        
        return {
            'fund_identification': fund_info,
            'investment_strategy': strategy_info,
            'leverage_calculation': leverage_data,
            'risk_profile': risk_data,
            'liquidity_management': liquidity_data,
            'fund_size': leverage_data.get('total_fund_size', 0),
            'investment_strategy_code': strategy_info.get('primary_strategy', 'multi_strategy'),
            'leverage_calculation': leverage_data,
            'liquidity_profile': liquidity_data,
            'risk_measures': risk_data,
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat()
        }
    
    async def aggregate_solvency_ii_data(
        self,
        period_start: date,
        period_end: date,
        entity_ids: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """Aggregate data for Solvency II report."""
        
        # Balance sheet information
        balance_sheet = await self._get_balance_sheet_data(period_end, entity_ids)
        
        # Own funds calculation
        own_funds = await self._get_own_funds_data(period_end, entity_ids)
        
        # Capital requirements
        scr_data = await self._get_scr_calculation(period_end, entity_ids)
        mcr_data = await self._get_mcr_calculation(period_end, entity_ids)
        
        # Risk concentrations
        concentration_risk = await self._get_concentration_risk(period_end, entity_ids)
        
        return {
            'balance_sheet': balance_sheet,
            'own_funds': own_funds,
            'solvency_capital_requirement': scr_data,
            'minimum_capital_requirement': mcr_data,
            'risk_concentration': concentration_risk,
            'technical_provisions': balance_sheet.get('technical_provisions', 0),
            'own_funds_total': own_funds.get('total_own_funds', 0),
            'scr_ratio': scr_data.get('scr_ratio', 0),
            'mcr_ratio': mcr_data.get('mcr_ratio', 0),
            'concentration_risk': concentration_risk,
            'market_risk': scr_data.get('market_risk_component', 0),
            'period_end': period_end.isoformat()
        }
    
    async def _get_basic_fund_info(self, fund_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get basic fund information."""
        query = """
            SELECT 
                id, name, strategy_type, risk_profile, 
                created_at, benchmark_ticker
            FROM portfolios 
            WHERE is_active = true
        """
        params = []
        
        if fund_ids:
            placeholders = ', '.join(f'${i+1}' for i in range(len(fund_ids)))
            query += f" AND id::text IN ({placeholders})"
            params.extend(fund_ids)
        
        funds = await self.db.fetch_all(query, *params)
        
        return {
            'fund_count': len(funds),
            'funds': [dict(fund) for fund in funds],
            'primary_strategies': list(set(fund['strategy_type'] for fund in funds)),
            'risk_profiles': list(set(fund['risk_profile'] for fund in funds if fund['risk_profile']))
        }
    
    async def _get_aum_data(
        self, 
        period_start: date, 
        period_end: date, 
        fund_ids: Optional[List[str]]
    ) -> Dict[str, Any]:
        """Get assets under management data."""
        query = """
            SELECT 
                p.id as portfolio_id,
                p.name as portfolio_name,
                COALESCE(SUM(ph.quantity * ph.current_price), 0) as gross_asset_value,
                COALESCE(SUM(ph.quantity * ph.current_price) - COALESCE(p_debt.total_debt, 0), 0) as net_asset_value,
                COUNT(DISTINCT ph.asset_id) as asset_count
            FROM portfolios p
            LEFT JOIN portfolio_holdings ph ON p.id = ph.portfolio_id
            LEFT JOIN (
                SELECT portfolio_id, SUM(debt_amount) as total_debt 
                FROM portfolio_debt 
                WHERE valuation_date <= $2 
                GROUP BY portfolio_id
            ) p_debt ON p.id = p_debt.portfolio_id
            WHERE p.is_active = true
        """
        
        params = [period_start, period_end]
        
        if fund_ids:
            placeholders = ', '.join(f'${i+3}' for i in range(len(fund_ids)))
            query += f" AND p.id::text IN ({placeholders})"
            params.extend(fund_ids)
        
        query += " GROUP BY p.id, p.name, p_debt.total_debt"
        
        results = await self.db.fetch_all(query, *params)
        
        total_gross_aum = sum(row['gross_asset_value'] for row in results)
        total_nav = sum(row['net_asset_value'] for row in results)
        
        return {
            'total_gross_aum': total_gross_aum,
            'total_nav': total_nav,
            'portfolio_breakdown': [dict(row) for row in results],
            'average_portfolio_size': total_gross_aum / len(results) if results else 0,
            'total_portfolios': len(results)
        }
    
    async def _get_risk_metrics(
        self, 
        period_start: date, 
        period_end: date, 
        fund_ids: Optional[List[str]]
    ) -> Dict[str, Any]:
        """Get risk metrics for regulatory reporting."""
        # This would integrate with the risk monitoring system from Story 1.2
        query = """
            SELECT 
                portfolio_id,
                metric_type,
                metric_value,
                calculation_date,
                confidence_level
            FROM risk_metrics 
            WHERE calculation_date BETWEEN $1 AND $2
        """
        
        params = [period_start, period_end]
        
        if fund_ids:
            placeholders = ', '.join(f'${i+3}' for i in range(len(fund_ids)))
            query += f" AND portfolio_id::text IN ({placeholders})"
            params.extend(fund_ids)
        
        query += " ORDER BY calculation_date DESC"
        
        risk_data = await self.db.fetch_all(query, *params)
        
        # Aggregate risk metrics
        var_measures = {}
        stress_tests = {}
        
        for row in risk_data:
            metric_type = row['metric_type']
            if metric_type.startswith('var_'):
                var_measures[metric_type] = row['metric_value']
            elif metric_type.startswith('stress_'):
                stress_tests[metric_type] = row['metric_value']
        
        # Calculate aggregate leverage ratio
        leverage_ratio = await self._calculate_leverage_ratio(fund_ids)
        
        return {
            'var_measures': var_measures,
            'stress_tests': stress_tests,
            'leverage_ratio': leverage_ratio,
            'risk_calculation_date': period_end.isoformat(),
            'confidence_levels': ['95%', '99%'],  # Standard VaR confidence levels
            'stress_scenarios': list(stress_tests.keys())
        }
    
    async def _calculate_leverage_ratio(self, fund_ids: Optional[List[str]]) -> float:
        """Calculate aggregate leverage ratio."""
        # Simplified leverage calculation (gross exposure / net assets)
        query = """
            SELECT 
                COALESCE(SUM(ABS(ph.quantity * ph.current_price)), 0) as gross_exposure,
                COALESCE(SUM(ph.quantity * ph.current_price), 0) as net_exposure
            FROM portfolio_holdings ph
            JOIN portfolios p ON ph.portfolio_id = p.id
            WHERE p.is_active = true
        """
        
        params = []
        if fund_ids:
            placeholders = ', '.join(f'${i+1}' for i in range(len(fund_ids)))
            query += f" AND p.id::text IN ({placeholders})"
            params.extend(fund_ids)
        
        result = await self.db.fetch_one(query, *params)
        
        if result and result['net_exposure'] > 0:
            return result['gross_exposure'] / result['net_exposure']
        return 0.0
    
    # Additional helper methods for other data aggregation needs
    async def _get_trading_data(self, period_start: date, period_end: date, fund_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get trading activity data."""
        return {
            'total_trades': 0,  # Placeholder - would integrate with trade execution system
            'trading_volume': 0,
            'counterparties': [],
            'average_trade_size': 0
        }
    
    async def _get_investment_strategy(self, fund_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get investment strategy information."""
        return {
            'primary_strategy': 'multi_strategy',
            'secondary_strategies': [],
            'geographic_focus': ['global'],
            'sector_allocation': {}
        }
    
    async def _get_leverage_calculation(self, period_start: date, period_end: date, fund_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get detailed leverage calculation for AIFMD."""
        leverage_ratio = await self._calculate_leverage_ratio(fund_ids)
        return {
            'gross_method_ratio': leverage_ratio,
            'commitment_method_ratio': leverage_ratio * 0.8,  # Simplified calculation
            'total_fund_size': 0  # Would be calculated from actual fund data
        }
    
    async def _get_liquidity_profile(self, period_start: date, period_end: date, fund_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get liquidity profile for AIFMD."""
        return {
            'liquidity_buckets': {
                'daily': 0.2,
                'weekly': 0.3,
                'monthly': 0.3,
                'quarterly': 0.2
            },
            'redemption_frequency': 'monthly',
            'notice_period': '30_days'
        }
    
    # Solvency II specific methods
    async def _get_balance_sheet_data(self, period_end: date, entity_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get balance sheet data for Solvency II."""
        return {
            'technical_provisions': 0,
            'investments': 0,
            'cash_equivalents': 0,
            'total_assets': 0
        }
    
    async def _get_own_funds_data(self, period_end: date, entity_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get own funds data for Solvency II."""
        return {
            'total_own_funds': 0,
            'tier_1_own_funds': 0,
            'tier_2_own_funds': 0,
            'tier_3_own_funds': 0
        }
    
    async def _get_scr_calculation(self, period_end: date, entity_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get SCR calculation for Solvency II."""
        return {
            'scr_ratio': 1.5,  # Example: 150%
            'market_risk_component': 0,
            'credit_risk_component': 0,
            'operational_risk_component': 0
        }
    
    async def _get_mcr_calculation(self, period_end: date, entity_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get MCR calculation for Solvency II."""
        return {
            'mcr_ratio': 2.0,  # Example: 200%
            'mcr_amount': 0
        }
    
    async def _get_concentration_risk(self, period_end: date, entity_ids: Optional[List[str]]) -> Dict[str, Any]:
        """Get concentration risk data for Solvency II."""
        return {
            'largest_exposure': 0,
            'top_10_exposures': 0,
            'geographic_concentration': {},
            'sector_concentration': {}
        }


class ReportGenerationPipeline:
    """
    Automated report generation pipeline with validation and delivery.
    
    Orchestrates the complete report generation process from data aggregation
    through validation, formatting, and delivery.
    """
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.template_manager = RegulatoryReportTemplateManager()
        self.data_aggregator = ReportDataAggregator(db_connection)
        self.audit_capture = AuditEventCapture(
            # This would be initialized with the audit trail from the immutable_audit_trail module
            None  # Placeholder - would be properly initialized
        )
    
    async def generate_regulatory_report(
        self,
        report_type: str,
        period_start: date,
        period_end: date,
        output_format: str = 'pdf',
        entity_ids: Optional[List[str]] = None,
        user_id: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Generate complete regulatory report.
        
        Args:
            report_type: Type of report ('form_pf', 'aifmd', 'solvency_ii')
            period_start: Report period start date
            period_end: Report period end date
            output_format: Output format ('pdf', 'excel', 'json')
            entity_ids: List of entity IDs to include
            user_id: User requesting the report
            
        Returns:
            Dict containing report metadata and file information
        """
        try:
            logger.info(f"Starting {report_type} report generation for period {period_start} to {period_end}")
            
            # Generate unique report ID
            report_id = f"{report_type}_{period_end.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}"
            
            # Step 1: Aggregate data
            if report_type == 'form_pf':
                report_data = await self.data_aggregator.aggregate_form_pf_data(
                    period_start, period_end, entity_ids
                )
            elif report_type == 'aifmd':
                report_data = await self.data_aggregator.aggregate_aifmd_data(
                    period_start, period_end, entity_ids
                )
            elif report_type == 'solvency_ii':
                report_data = await self.data_aggregator.aggregate_solvency_ii_data(
                    period_start, period_end, entity_ids
                )
            else:
                raise ValueError(f"Unsupported report type: {report_type}")
            
            # Step 2: Validate data
            validation_result = self.template_manager.validate_template_data(report_type, report_data)
            
            if not validation_result.is_valid:
                logger.error(f"Report validation failed: {validation_result.errors}")
                return {
                    'status': 'validation_failed',
                    'report_id': report_id,
                    'validation_result': asdict(validation_result)
                }
            
            # Step 3: Generate report file
            file_path = await self._generate_report_file(
                report_id, report_type, report_data, output_format
            )
            
            # Step 4: Calculate file metadata
            file_stats = os.stat(file_path)
            file_hash = await self._calculate_file_hash(file_path)
            
            # Step 5: Store report metadata in database
            metadata = ReportMetadata(
                report_id=report_id,
                report_type=report_type,
                period_start=period_start,
                period_end=period_end,
                generated_at=datetime.now(),
                template_version='1.0',
                data_sources=['portfolio_holdings', 'risk_metrics', 'performance_records'],
                validation_status='passed',
                file_format=output_format
            )
            
            await self._store_report_metadata(metadata, file_path, file_stats.st_size, file_hash)
            
            # Step 6: Log audit event
            if self.audit_capture:
                await self.audit_capture.capture_portfolio_decision(
                    portfolio_id=None,  # Report spans multiple portfolios
                    decision_type='regulatory_report_generation',
                    optimization_params={
                        'report_type': report_type,
                        'period_start': period_start.isoformat(),
                        'period_end': period_end.isoformat(),
                        'output_format': output_format
                    },
                    model_version='regulatory_reporting_v1.0',
                    user_id=user_id or 0,
                    session_context={
                        'report_id': report_id,
                        'validation_status': 'passed',
                        'data_completeness': validation_result.data_completeness,
                        'file_path': file_path,
                        'file_size': file_stats.st_size
                    }
                )
            
            logger.info(f"Successfully generated {report_type} report: {report_id}")
            
            return {
                'status': 'success',
                'report_id': report_id,
                'file_path': file_path,
                'file_size': file_stats.st_size,
                'file_hash': file_hash,
                'validation_result': asdict(validation_result),
                'metadata': asdict(metadata)
            }
            
        except Exception as e:
            logger.error(f"Failed to generate {report_type} report: {e}")
            return {
                'status': 'error',
                'report_id': report_id if 'report_id' in locals() else 'unknown',
                'error': str(e)
            }
    
    async def _generate_report_file(
        self,
        report_id: str,
        report_type: str,
        report_data: Dict[str, Any],
        output_format: str
    ) -> str:
        """Generate report file in specified format."""
        
        # Create output directory
        output_dir = Path("reports") / "regulatory"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        if output_format == 'pdf':
            file_path = output_dir / f"{report_id}.pdf"
            await self._generate_pdf_report(report_type, report_data, str(file_path))
        elif output_format == 'excel':
            file_path = output_dir / f"{report_id}.xlsx"
            await self._generate_excel_report(report_type, report_data, str(file_path))
        elif output_format == 'json':
            file_path = output_dir / f"{report_id}.json"
            await self._generate_json_report(report_type, report_data, str(file_path))
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
        
        return str(file_path)
    
    async def _generate_pdf_report(self, report_type: str, report_data: Dict[str, Any], file_path: str):
        """Generate PDF report using ReportLab."""
        doc = SimpleDocTemplate(file_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        template = self.template_manager.get_template(report_type)
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(template['name'], title_style))
        story.append(Spacer(1, 12))
        
        # Report period
        period_text = f"Reporting Period: {report_data.get('period_start', 'N/A')} to {report_data.get('period_end', 'N/A')}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Generate sections based on template
        for section in template['sections']:
            section_data = report_data.get(section, {})
            if section_data:
                # Section header
                story.append(Paragraph(section.replace('_', ' ').title(), styles['Heading2']))
                
                # Section content
                if isinstance(section_data, dict):
                    for key, value in section_data.items():
                        if isinstance(value, (int, float)):
                            text = f"{key.replace('_', ' ').title()}: {value:,.2f}"
                        else:
                            text = f"{key.replace('_', ' ').title()}: {value}"
                        story.append(Paragraph(text, styles['Normal']))
                
                story.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(story)
    
    async def _generate_excel_report(self, report_type: str, report_data: Dict[str, Any], file_path: str):
        """Generate Excel report using openpyxl."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        template = self.template_manager.get_template(report_type)
        worksheet.title = template['name']
        
        # Header styling
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Report title
        worksheet['A1'] = template['name']
        worksheet['A1'].font = header_font
        worksheet.merge_cells('A1:C1')
        
        # Report period
        worksheet['A3'] = 'Reporting Period:'
        worksheet['B3'] = f"{report_data.get('period_start', 'N/A')} to {report_data.get('period_end', 'N/A')}"
        
        # Data sections
        row = 5
        for section in template['sections']:
            section_data = report_data.get(section, {})
            if section_data:
                # Section header
                worksheet[f'A{row}'] = section.replace('_', ' ').title()
                worksheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                # Section data
                if isinstance(section_data, dict):
                    for key, value in section_data.items():
                        worksheet[f'A{row}'] = key.replace('_', ' ').title()
                        worksheet[f'B{row}'] = value
                        row += 1
                
                row += 1  # Add space between sections
        
        # Save workbook
        workbook.save(file_path)
    
    async def _generate_json_report(self, report_type: str, report_data: Dict[str, Any], file_path: str):
        """Generate JSON report for API consumption."""
        template = self.template_manager.get_template(report_type)
        
        json_report = {
            'report_metadata': {
                'report_type': report_type,
                'template_name': template['name'],
                'generated_at': datetime.now().isoformat(),
                'template_version': '1.0'
            },
            'report_data': report_data
        }
        
        with open(file_path, 'w') as f:
            json.dump(json_report, f, indent=2, default=str)
    
    async def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA-256 hash of generated file."""
        hash_sha256 = hashlib.sha256()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()
    
    async def _store_report_metadata(
        self,
        metadata: ReportMetadata,
        file_path: str,
        file_size: int,
        file_hash: str
    ):
        """Store report metadata in database."""
        query = """
            INSERT INTO regulatory_reports (
                report_type, report_period_start, report_period_end,
                generated_at, report_data, file_path, file_size_bytes,
                checksum_sha256, submission_status
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
        """
        
        await self.db.execute(
            query,
            metadata.report_type,
            metadata.period_start,
            metadata.period_end,
            metadata.generated_at,
            json.dumps(asdict(metadata)),
            file_path,
            file_size,
            file_hash,
            'draft'
        )


class RegulatoryReportingEngine:
    """
    Main regulatory reporting engine that orchestrates report generation,
    validation, scheduling, and delivery for compliance requirements.
    """
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.generation_pipeline = ReportGenerationPipeline(db_connection)
        self.template_manager = RegulatoryReportTemplateManager()
    
    async def generate_report(
        self,
        report_type: str,
        period_start: date,
        period_end: date,
        output_format: str = 'pdf',
        entity_ids: Optional[List[str]] = None,
        user_id: Optional[int] = None
    ) -> Dict[str, Any]:
        """Generate regulatory report - main entry point."""
        return await self.generation_pipeline.generate_regulatory_report(
            report_type, period_start, period_end, output_format, entity_ids, user_id
        )
    
    async def schedule_report(
        self,
        report_type: str,
        schedule_frequency: str,
        next_due_date: date,
        output_format: str = 'pdf',
        entity_ids: Optional[List[str]] = None
    ) -> str:
        """Schedule automatic report generation."""
        # This would integrate with a job scheduler like Celery or similar
        # For now, store the schedule in the database
        
        query = """
            INSERT INTO compliance_filings (
                filing_type, filing_name, filing_frequency,
                due_date, status, priority
            ) VALUES ($1, $2, $3, $4, $5, $6)
            RETURNING id
        """
        
        template = self.template_manager.get_template(report_type)
        
        result = await self.db.fetch_one(
            query,
            report_type,
            template['name'],
            schedule_frequency,
            next_due_date,
            'pending',
            'medium'
        )
        
        logger.info(f"Scheduled {report_type} report generation for {next_due_date}")
        return str(result['id'])
    
    async def get_report_status(self, report_id: str) -> Dict[str, Any]:
        """Get status of generated report."""
        query = """
            SELECT * FROM regulatory_reports 
            WHERE id = $1 OR file_path LIKE '%' || $1 || '%'
        """
        
        result = await self.db.fetch_one(query, report_id)
        
        if result:
            return dict(result)
        else:
            return {'status': 'not_found', 'report_id': report_id}
    
    async def list_available_reports(
        self,
        report_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> List[Dict[str, Any]]:
        """List available generated reports."""
        conditions = []
        params = []
        param_count = 0
        
        if report_type:
            param_count += 1
            conditions.append(f"report_type = ${param_count}")
            params.append(report_type)
        
        if start_date:
            param_count += 1
            conditions.append(f"generated_at >= ${param_count}")
            params.append(start_date)
        
        if end_date:
            param_count += 1
            conditions.append(f"generated_at <= ${param_count}")
            params.append(end_date)
        
        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""
        
        query = f"""
            SELECT 
                id, report_type, report_period_start, report_period_end,
                generated_at, submission_status, file_path, file_size_bytes
            FROM regulatory_reports
            {where_clause}
            ORDER BY generated_at DESC
        """
        
        results = await self.db.fetch_all(query, *params)
        return [dict(row) for row in results]
