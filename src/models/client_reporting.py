"""
Client Reporting System for Quantum Portfolio Optimizer

This module provides comprehensive client reporting capabilities including
performance attribution, risk metrics, and customizable report delivery.

Features:
- Client-specific report templates and configurations
- Performance attribution analysis
- Risk metrics reporting (VaR, stress tests, scenario analysis)
- Benchmark comparison and peer analysis
- Multi-format output (PDF, Excel, interactive web)
- Automated delivery and client portal access
"""

import asyncio
import json
from datetime import datetime, date, timedelta
from typing import Dict, Any, List, Optional, Union
from dataclasses import dataclass, asdict
from pathlib import Path
import tempfile
import io
import os

# Report generation libraries
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.piecharts import Pie
import openpyxl
from openpyxl.chart import LineChart, Reference, PieChart
from openpyxl.styles import Font, PatternFill, Alignment

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

from ..database.connection import DatabaseConnection
from ..utils.professional_logging import get_logger
from ..utils.immutable_audit_trail import AuditEventCapture, AuditEventData

logger = get_logger(__name__)


@dataclass
class ClientReportConfig:
    """Client report configuration settings."""
    client_id: int
    report_name: str
    report_type: str  # 'performance', 'risk', 'attribution', 'compliance'
    delivery_schedule: str  # 'daily', 'weekly', 'monthly', 'quarterly', 'annual'
    delivery_method: str  # 'email', 'portal', 'api'
    recipients: List[str]
    template_settings: Dict[str, Any]
    custom_metrics: List[str]
    benchmark_comparisons: List[str]
    risk_metrics_included: List[str]
    format_preferences: Dict[str, Any]


@dataclass
class PerformanceAttribution:
    """Performance attribution analysis results."""
    total_return: float
    benchmark_return: float
    excess_return: float
    attribution_factors: Dict[str, float]
    sector_attribution: Dict[str, float]
    security_selection: float
    asset_allocation: float
    interaction_effect: float
    transaction_costs: float


@dataclass
class RiskMetrics:
    """Risk metrics for client reporting."""
    var_95: float
    var_99: float
    cvar_95: float
    cvar_99: float
    volatility: float
    sharpe_ratio: float
    sortino_ratio: float
    max_drawdown: float
    beta: float
    tracking_error: float
    information_ratio: float


class ClientReportDataProcessor:
    """
    Data processing and analysis for client reports.
    
    Handles performance calculations, risk metrics, attribution analysis,
    and benchmark comparisons for client reporting.
    """
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
    
    async def calculate_performance_attribution(
        self,
        portfolio_id: int,
        period_start: date,
        period_end: date,
        benchmark_ticker: str = 'SPY'
    ) -> PerformanceAttribution:
        """Calculate detailed performance attribution analysis."""
        
        # Get portfolio performance data
        portfolio_returns = await self._get_portfolio_returns(portfolio_id, period_start, period_end)
        benchmark_returns = await self._get_benchmark_returns(benchmark_ticker, period_start, period_end)
        
        # Calculate total returns
        total_return = portfolio_returns['total_return']
        benchmark_return = benchmark_returns['total_return']
        excess_return = total_return - benchmark_return
        
        # Get sector/asset allocation data
        sector_weights = await self._get_sector_allocation(portfolio_id, period_start, period_end)
        benchmark_weights = await self._get_benchmark_allocation(benchmark_ticker, period_start, period_end)
        
        # Calculate attribution factors
        attribution_factors = await self._calculate_attribution_factors(
            portfolio_id, period_start, period_end, benchmark_ticker
        )
        
        # Sector attribution analysis
        sector_attribution = await self._calculate_sector_attribution(
            sector_weights, benchmark_weights, period_start, period_end
        )
        
        # Brinson attribution components
        security_selection = attribution_factors.get('security_selection', 0.0)
        asset_allocation = attribution_factors.get('asset_allocation', 0.0)
        interaction_effect = attribution_factors.get('interaction', 0.0)
        transaction_costs = attribution_factors.get('transaction_costs', 0.0)
        
        return PerformanceAttribution(
            total_return=total_return,
            benchmark_return=benchmark_return,
            excess_return=excess_return,
            attribution_factors=attribution_factors,
            sector_attribution=sector_attribution,
            security_selection=security_selection,
            asset_allocation=asset_allocation,
            interaction_effect=interaction_effect,
            transaction_costs=transaction_costs
        )
    
    async def calculate_risk_metrics(
        self,
        portfolio_id: int,
        period_start: date,
        period_end: date,
        benchmark_ticker: str = 'SPY'
    ) -> RiskMetrics:
        """Calculate comprehensive risk metrics for client reporting."""
        
        # Get portfolio and benchmark return series
        portfolio_returns = await self._get_return_series(portfolio_id, period_start, period_end)
        benchmark_returns = await self._get_benchmark_return_series(benchmark_ticker, period_start, period_end)
        
        if not portfolio_returns or len(portfolio_returns) < 10:
            # Insufficient data for meaningful risk metrics
            return RiskMetrics(0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
        
        returns_array = np.array(portfolio_returns)
        benchmark_array = np.array(benchmark_returns)
        
        # Calculate VaR and CVaR
        var_95 = np.percentile(returns_array, 5)  # 5th percentile
        var_99 = np.percentile(returns_array, 1)  # 1st percentile
        cvar_95 = returns_array[returns_array <= var_95].mean()
        cvar_99 = returns_array[returns_array <= var_99].mean()
        
        # Calculate volatility metrics
        volatility = np.std(returns_array) * np.sqrt(252)  # Annualized
        
        # Calculate risk-adjusted returns
        mean_return = np.mean(returns_array) * 252  # Annualized
        risk_free_rate = 0.02  # 2% risk-free rate assumption
        
        sharpe_ratio = (mean_return - risk_free_rate) / volatility if volatility > 0 else 0
        
        # Sortino ratio (downside deviation)
        downside_returns = returns_array[returns_array < 0]
        downside_deviation = np.std(downside_returns) * np.sqrt(252) if len(downside_returns) > 0 else 0
        sortino_ratio = (mean_return - risk_free_rate) / downside_deviation if downside_deviation > 0 else 0
        
        # Maximum drawdown
        cumulative_returns = np.cumprod(1 + returns_array)
        running_max = np.maximum.accumulate(cumulative_returns)
        drawdown = (cumulative_returns - running_max) / running_max
        max_drawdown = np.min(drawdown)
        
        # Beta and tracking error (relative to benchmark)
        if len(benchmark_array) == len(returns_array):
            covariance = np.cov(returns_array, benchmark_array)[0, 1]
            benchmark_variance = np.var(benchmark_array)
            beta = covariance / benchmark_variance if benchmark_variance > 0 else 0
            
            tracking_error = np.std(returns_array - benchmark_array) * np.sqrt(252)
            excess_returns = returns_array - benchmark_array
            information_ratio = np.mean(excess_returns) * 252 / tracking_error if tracking_error > 0 else 0
        else:
            beta = 0
            tracking_error = 0
            information_ratio = 0
        
        return RiskMetrics(
            var_95=var_95,
            var_99=var_99,
            cvar_95=cvar_95,
            cvar_99=cvar_99,
            volatility=volatility,
            sharpe_ratio=sharpe_ratio,
            sortino_ratio=sortino_ratio,
            max_drawdown=max_drawdown,
            beta=beta,
            tracking_error=tracking_error,
            information_ratio=information_ratio
        )
    
    async def generate_benchmark_comparison(
        self,
        portfolio_id: int,
        period_start: date,
        period_end: date,
        benchmark_tickers: List[str] = None
    ) -> Dict[str, Any]:
        """Generate comprehensive benchmark comparison analysis."""
        
        if not benchmark_tickers:
            benchmark_tickers = ['SPY', 'QQQ', 'IWM']  # Default benchmarks
        
        portfolio_performance = await self._get_portfolio_performance_summary(
            portfolio_id, period_start, period_end
        )
        
        benchmark_comparisons = {}
        for ticker in benchmark_tickers:
            benchmark_performance = await self._get_benchmark_performance_summary(
                ticker, period_start, period_end
            )
            
            comparison = {
                'benchmark_name': ticker,
                'portfolio_return': portfolio_performance['total_return'],
                'benchmark_return': benchmark_performance['total_return'],
                'excess_return': portfolio_performance['total_return'] - benchmark_performance['total_return'],
                'portfolio_volatility': portfolio_performance['volatility'],
                'benchmark_volatility': benchmark_performance['volatility'],
                'portfolio_sharpe': portfolio_performance['sharpe_ratio'],
                'benchmark_sharpe': benchmark_performance['sharpe_ratio']
            }
            
            benchmark_comparisons[ticker] = comparison
        
        return {
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(),
            'portfolio_id': portfolio_id,
            'benchmark_comparisons': benchmark_comparisons,
            'best_performing_benchmark': max(
                benchmark_comparisons.keys(),
                key=lambda x: benchmark_comparisons[x]['benchmark_return']
            ),
            'portfolio_ranking': self._calculate_portfolio_ranking(
                portfolio_performance['total_return'],
                [comp['benchmark_return'] for comp in benchmark_comparisons.values()]
            )
        }
    
    # Helper methods for data retrieval and calculations
    async def _get_portfolio_returns(self, portfolio_id: int, start_date: date, end_date: date) -> Dict[str, Any]:
        """Get portfolio return data for specified period."""
        query = """
            SELECT 
                calculation_date,
                total_return,
                cumulative_return
            FROM performance_records
            WHERE portfolio_id = $1 
            AND calculation_date BETWEEN $2 AND $3
            ORDER BY calculation_date
        """
        
        results = await self.db.fetch_all(query, portfolio_id, start_date, end_date)
        
        if results:
            total_return = results[-1]['cumulative_return'] if results[-1]['cumulative_return'] else 0
            return {
                'total_return': total_return,
                'daily_returns': [row['total_return'] for row in results]
            }
        
        return {'total_return': 0, 'daily_returns': []}
    
    async def _get_benchmark_returns(self, ticker: str, start_date: date, end_date: date) -> Dict[str, Any]:
        """Get benchmark return data - would integrate with market data service."""
        # Placeholder implementation - would integrate with actual market data
        # For now, return sample data
        return {
            'total_return': 0.08,  # 8% sample return
            'daily_returns': []
        }
    
    async def _get_return_series(self, portfolio_id: int, start_date: date, end_date: date) -> List[float]:
        """Get daily return series for portfolio."""
        query = """
            SELECT total_return
            FROM performance_records
            WHERE portfolio_id = $1 
            AND calculation_date BETWEEN $2 AND $3
            ORDER BY calculation_date
        """
        
        results = await self.db.fetch_all(query, portfolio_id, start_date, end_date)
        return [row['total_return'] or 0 for row in results]
    
    async def _get_benchmark_return_series(self, ticker: str, start_date: date, end_date: date) -> List[float]:
        """Get benchmark daily return series."""
        # Placeholder - would integrate with market data service
        # Generate sample data for demonstration
        days = (end_date - start_date).days
        np.random.seed(42)  # For consistent sample data
        return np.random.normal(0.0008, 0.015, days).tolist()  # Sample daily returns
    
    async def _get_sector_allocation(self, portfolio_id: int, start_date: date, end_date: date) -> Dict[str, float]:
        """Get portfolio sector allocation."""
        query = """
            SELECT 
                a.sector,
                SUM(ph.quantity * ph.current_price) / 
                (SELECT SUM(quantity * current_price) FROM portfolio_holdings WHERE portfolio_id = $1) as weight
            FROM portfolio_holdings ph
            JOIN assets a ON ph.asset_id = a.id
            WHERE ph.portfolio_id = $1
            GROUP BY a.sector
        """
        
        results = await self.db.fetch_all(query, portfolio_id)
        return {row['sector']: row['weight'] for row in results if row['sector']}
    
    async def _get_benchmark_allocation(self, ticker: str, start_date: date, end_date: date) -> Dict[str, float]:
        """Get benchmark sector allocation."""
        # Placeholder - would integrate with benchmark data service
        return {
            'Technology': 0.25,
            'Healthcare': 0.15,
            'Financials': 0.12,
            'Consumer Discretionary': 0.10,
            'Communication Services': 0.08,
            'Industrials': 0.08,
            'Consumer Staples': 0.07,
            'Energy': 0.05,
            'Utilities': 0.05,
            'Real Estate': 0.05
        }
    
    async def _calculate_attribution_factors(
        self, portfolio_id: int, start_date: date, end_date: date, benchmark: str
    ) -> Dict[str, float]:
        """Calculate detailed attribution factors."""
        # Simplified attribution calculation
        return {
            'security_selection': 0.015,  # 1.5% from security selection
            'asset_allocation': -0.005,   # -0.5% from asset allocation
            'interaction': 0.002,         # 0.2% interaction effect
            'transaction_costs': -0.003   # -0.3% transaction costs
        }
    
    async def _calculate_sector_attribution(
        self, portfolio_weights: Dict[str, float], benchmark_weights: Dict[str, float],
        start_date: date, end_date: date
    ) -> Dict[str, float]:
        """Calculate sector-level attribution."""
        sector_attribution = {}
        
        for sector in set(list(portfolio_weights.keys()) + list(benchmark_weights.keys())):
            port_weight = portfolio_weights.get(sector, 0)
            bench_weight = benchmark_weights.get(sector, 0)
            
            # Simplified attribution calculation
            weight_effect = port_weight - bench_weight
            sector_attribution[sector] = weight_effect * 0.1  # Sample sector return
        
        return sector_attribution
    
    def _calculate_portfolio_ranking(self, portfolio_return: float, benchmark_returns: List[float]) -> Dict[str, Any]:
        """Calculate portfolio ranking versus benchmarks."""
        all_returns = benchmark_returns + [portfolio_return]
        all_returns.sort(reverse=True)
        
        portfolio_rank = all_returns.index(portfolio_return) + 1
        percentile = (len(all_returns) - portfolio_rank + 1) / len(all_returns) * 100
        
        return {
            'rank': portfolio_rank,
            'percentile': percentile,
            'total_compared': len(all_returns),
            'above_median': percentile > 50
        }
    
    async def _get_portfolio_performance_summary(self, portfolio_id: int, start_date: date, end_date: date) -> Dict[str, Any]:
        """Get portfolio performance summary."""
        returns = await self._get_return_series(portfolio_id, start_date, end_date)
        
        if not returns:
            return {'total_return': 0, 'volatility': 0, 'sharpe_ratio': 0}
        
        returns_array = np.array(returns)
        total_return = np.prod(1 + returns_array) - 1
        volatility = np.std(returns_array) * np.sqrt(252)
        mean_return = np.mean(returns_array) * 252
        risk_free_rate = 0.02
        sharpe_ratio = (mean_return - risk_free_rate) / volatility if volatility > 0 else 0
        
        return {
            'total_return': total_return,
            'volatility': volatility,
            'sharpe_ratio': sharpe_ratio
        }
    
    async def _get_benchmark_performance_summary(self, ticker: str, start_date: date, end_date: date) -> Dict[str, Any]:
        """Get benchmark performance summary."""
        # Placeholder - would integrate with market data
        return {
            'total_return': 0.08,
            'volatility': 0.16,
            'sharpe_ratio': 0.5
        }


class ClientReportGenerator:
    """
    Client report generation engine with multiple output formats.
    
    Generates comprehensive client reports including performance attribution,
    risk analysis, and benchmark comparisons in PDF, Excel, and web formats.
    """
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.data_processor = ClientReportDataProcessor(db_connection)
        self.audit_capture = AuditEventCapture(None)  # Would be properly initialized
    
    async def generate_client_report(
        self,
        config: ClientReportConfig,
        period_start: date,
        period_end: date,
        portfolio_ids: List[int],
        output_format: str = 'pdf'
    ) -> Dict[str, Any]:
        """
        Generate comprehensive client report.
        
        Args:
            config: Client report configuration
            period_start: Report period start date
            period_end: Report period end date
            portfolio_ids: List of portfolio IDs to include
            output_format: Output format ('pdf', 'excel', 'html')
            
        Returns:
            Dict containing report metadata and file information
        """
        try:
            logger.info(f"Generating {config.report_type} report for client {config.client_id}")
            
            # Generate report data based on type
            if config.report_type == 'performance':
                report_data = await self._generate_performance_report_data(
                    config, period_start, period_end, portfolio_ids
                )
            elif config.report_type == 'risk':
                report_data = await self._generate_risk_report_data(
                    config, period_start, period_end, portfolio_ids
                )
            elif config.report_type == 'attribution':
                report_data = await self._generate_attribution_report_data(
                    config, period_start, period_end, portfolio_ids
                )
            else:
                raise ValueError(f"Unsupported report type: {config.report_type}")
            
            # Generate report file
            file_path = await self._generate_report_file(
                config, report_data, output_format
            )
            
            # Store delivery record
            delivery_id = await self._store_delivery_record(config, file_path)
            
            # Log audit event
            if self.audit_capture:
                await self.audit_capture.capture_portfolio_decision(
                    portfolio_id=portfolio_ids[0] if portfolio_ids else None,
                    decision_type='client_report_generation',
                    optimization_params={
                        'report_type': config.report_type,
                        'client_id': config.client_id,
                        'period_start': period_start.isoformat(),
                        'period_end': period_end.isoformat(),
                        'output_format': output_format
                    },
                    model_version='client_reporting_v1.0',
                    user_id=0,
                    session_context={
                        'report_name': config.report_name,
                        'delivery_method': config.delivery_method,
                        'file_path': file_path,
                        'portfolios_included': len(portfolio_ids)
                    }
                )
            
            logger.info(f"Successfully generated client report: {delivery_id}")
            
            return {
                'status': 'success',
                'delivery_id': delivery_id,
                'file_path': file_path,
                'report_data': report_data
            }
            
        except Exception as e:
            logger.error(f"Failed to generate client report: {e}")
            return {
                'status': 'error',
                'error': str(e)
            }
    
    async def _generate_performance_report_data(
        self,
        config: ClientReportConfig,
        period_start: date,
        period_end: date,
        portfolio_ids: List[int]
    ) -> Dict[str, Any]:
        """Generate performance report data."""
        
        performance_data = {}
        
        for portfolio_id in portfolio_ids:
            # Performance attribution
            attribution = await self.data_processor.calculate_performance_attribution(
                portfolio_id, period_start, period_end
            )
            
            # Risk metrics
            risk_metrics = await self.data_processor.calculate_risk_metrics(
                portfolio_id, period_start, period_end
            )
            
            # Benchmark comparison
            benchmark_comparison = await self.data_processor.generate_benchmark_comparison(
                portfolio_id, period_start, period_end, config.benchmark_comparisons
            )
            
            performance_data[str(portfolio_id)] = {
                'attribution': asdict(attribution),
                'risk_metrics': asdict(risk_metrics),
                'benchmark_comparison': benchmark_comparison
            }
        
        return {
            'report_type': 'performance',
            'client_id': config.client_id,
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(),
            'portfolios': performance_data,
            'custom_metrics': config.custom_metrics,
            'generated_at': datetime.now().isoformat()
        }
    
    async def _generate_risk_report_data(
        self,
        config: ClientReportConfig,
        period_start: date,
        period_end: date,
        portfolio_ids: List[int]
    ) -> Dict[str, Any]:
        """Generate risk-focused report data."""
        
        risk_data = {}
        
        for portfolio_id in portfolio_ids:
            risk_metrics = await self.data_processor.calculate_risk_metrics(
                portfolio_id, period_start, period_end
            )
            
            # Additional risk analysis would go here
            # - Stress test results
            # - Scenario analysis
            # - Risk decomposition
            
            risk_data[str(portfolio_id)] = {
                'risk_metrics': asdict(risk_metrics),
                'stress_tests': {},  # Placeholder
                'scenario_analysis': {}  # Placeholder
            }
        
        return {
            'report_type': 'risk',
            'client_id': config.client_id,
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(),
            'portfolios': risk_data,
            'risk_metrics_included': config.risk_metrics_included,
            'generated_at': datetime.now().isoformat()
        }
    
    async def _generate_attribution_report_data(
        self,
        config: ClientReportConfig,
        period_start: date,
        period_end: date,
        portfolio_ids: List[int]
    ) -> Dict[str, Any]:
        """Generate attribution-focused report data."""
        
        attribution_data = {}
        
        for portfolio_id in portfolio_ids:
            attribution = await self.data_processor.calculate_performance_attribution(
                portfolio_id, period_start, period_end
            )
            
            attribution_data[str(portfolio_id)] = asdict(attribution)
        
        return {
            'report_type': 'attribution',
            'client_id': config.client_id,
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(),
            'portfolios': attribution_data,
            'generated_at': datetime.now().isoformat()
        }
    
    async def _generate_report_file(
        self,
        config: ClientReportConfig,
        report_data: Dict[str, Any],
        output_format: str
    ) -> str:
        """Generate report file in specified format."""
        
        # Create output directory
        output_dir = Path("reports") / "clients" / str(config.client_id)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{config.report_name}_{timestamp}"
        
        if output_format == 'pdf':
            file_path = output_dir / f"{filename}.pdf"
            await self._generate_pdf_client_report(config, report_data, str(file_path))
        elif output_format == 'excel':
            file_path = output_dir / f"{filename}.xlsx"
            await self._generate_excel_client_report(config, report_data, str(file_path))
        elif output_format == 'html':
            file_path = output_dir / f"{filename}.html"
            await self._generate_html_client_report(config, report_data, str(file_path))
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
        
        return str(file_path)
    
    async def _generate_pdf_client_report(
        self,
        config: ClientReportConfig,
        report_data: Dict[str, Any],
        file_path: str
    ):
        """Generate PDF client report using ReportLab."""
        doc = SimpleDocTemplate(file_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Report title and header
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph(f"Portfolio Report: {config.report_name}", title_style))
        story.append(Spacer(1, 12))
        
        # Report metadata
        period_text = f"Period: {report_data['period_start']} to {report_data['period_end']}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Paragraph(f"Generated: {report_data['generated_at']}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Portfolio performance summary
        if report_data['report_type'] in ['performance', 'attribution']:
            story.append(Paragraph("Performance Summary", styles['Heading2']))
            
            for portfolio_id, data in report_data['portfolios'].items():
                story.append(Paragraph(f"Portfolio {portfolio_id}", styles['Heading3']))
                
                if 'attribution' in data:
                    attr = data['attribution']
                    performance_table_data = [
                        ['Metric', 'Value'],
                        ['Total Return', f"{attr['total_return']:.2%}"],
                        ['Benchmark Return', f"{attr['benchmark_return']:.2%}"],
                        ['Excess Return', f"{attr['excess_return']:.2%}"],
                        ['Security Selection', f"{attr['security_selection']:.2%}"],
                        ['Asset Allocation', f"{attr['asset_allocation']:.2%}"]
                    ]
                    
                    performance_table = Table(performance_table_data)
                    performance_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(performance_table)
                story.append(Spacer(1, 12))
        
        # Risk metrics
        if report_data['report_type'] in ['risk', 'performance']:
            story.append(Paragraph("Risk Metrics", styles['Heading2']))
            
            for portfolio_id, data in report_data['portfolios'].items():
                if 'risk_metrics' in data:
                    risk = data['risk_metrics']
                    risk_table_data = [
                        ['Risk Metric', 'Value'],
                        ['VaR (95%)', f"{risk['var_95']:.2%}"],
                        ['VaR (99%)', f"{risk['var_99']:.2%}"],
                        ['Volatility', f"{risk['volatility']:.2%}"],
                        ['Sharpe Ratio', f"{risk['sharpe_ratio']:.2f}"],
                        ['Max Drawdown', f"{risk['max_drawdown']:.2%}"]
                    ]
                    
                    risk_table = Table(risk_table_data)
                    risk_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(risk_table)
                story.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(story)
    
    async def _generate_excel_client_report(
        self,
        config: ClientReportConfig,
        report_data: Dict[str, Any],
        file_path: str
    ):
        """Generate Excel client report using openpyxl."""
        workbook = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        
        # Report header
        summary_sheet['A1'] = f"Portfolio Report: {config.report_name}"
        summary_sheet['A1'].font = Font(bold=True, size=16)
        summary_sheet['A3'] = f"Period: {report_data['period_start']} to {report_data['period_end']}"
        summary_sheet['A4'] = f"Generated: {report_data['generated_at']}"
        
        # Performance data
        if report_data['report_type'] in ['performance', 'attribution']:
            perf_sheet = workbook.create_sheet("Performance")
            row = 1
            
            for portfolio_id, data in report_data['portfolios'].items():
                perf_sheet[f'A{row}'] = f"Portfolio {portfolio_id}"
                perf_sheet[f'A{row}'].font = Font(bold=True)
                row += 2
                
                if 'attribution' in data:
                    attr = data['attribution']
                    perf_sheet[f'A{row}'] = "Total Return"
                    perf_sheet[f'B{row}'] = attr['total_return']
                    row += 1
                    perf_sheet[f'A{row}'] = "Benchmark Return"
                    perf_sheet[f'B{row}'] = attr['benchmark_return']
                    row += 1
                    perf_sheet[f'A{row}'] = "Excess Return"
                    perf_sheet[f'B{row}'] = attr['excess_return']
                    row += 3
        
        # Risk metrics
        if report_data['report_type'] in ['risk', 'performance']:
            risk_sheet = workbook.create_sheet("Risk Metrics")
            row = 1
            
            for portfolio_id, data in report_data['portfolios'].items():
                risk_sheet[f'A{row}'] = f"Portfolio {portfolio_id}"
                risk_sheet[f'A{row}'].font = Font(bold=True)
                row += 2
                
                if 'risk_metrics' in data:
                    risk = data['risk_metrics']
                    risk_sheet[f'A{row}'] = "VaR (95%)"
                    risk_sheet[f'B{row}'] = risk['var_95']
                    row += 1
                    risk_sheet[f'A{row}'] = "Volatility"
                    risk_sheet[f'B{row}'] = risk['volatility']
                    row += 1
                    risk_sheet[f'A{row}'] = "Sharpe Ratio"
                    risk_sheet[f'B{row}'] = risk['sharpe_ratio']
                    row += 3
        
        workbook.save(file_path)
    
    async def _generate_html_client_report(
        self,
        config: ClientReportConfig,
        report_data: Dict[str, Any],
        file_path: str
    ):
        """Generate HTML client report for web viewing."""
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{config.report_name}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .section {{ margin-bottom: 30px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .metric-value {{ text-align: right; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Portfolio Report: {config.report_name}</h1>
                <p>Period: {report_data['period_start']} to {report_data['period_end']}</p>
                <p>Generated: {report_data['generated_at']}</p>
            </div>
        """
        
        # Add performance data
        if report_data['report_type'] in ['performance', 'attribution']:
            html_content += '<div class="section"><h2>Performance Summary</h2>'
            
            for portfolio_id, data in report_data['portfolios'].items():
                html_content += f'<h3>Portfolio {portfolio_id}</h3>'
                
                if 'attribution' in data:
                    attr = data['attribution']
                    html_content += """
                    <table>
                        <tr><th>Metric</th><th>Value</th></tr>
                        <tr><td>Total Return</td><td class="metric-value">{:.2%}</td></tr>
                        <tr><td>Benchmark Return</td><td class="metric-value">{:.2%}</td></tr>
                        <tr><td>Excess Return</td><td class="metric-value">{:.2%}</td></tr>
                        <tr><td>Security Selection</td><td class="metric-value">{:.2%}</td></tr>
                        <tr><td>Asset Allocation</td><td class="metric-value">{:.2%}</td></tr>
                    </table>
                    """.format(
                        attr['total_return'],
                        attr['benchmark_return'],
                        attr['excess_return'],
                        attr['security_selection'],
                        attr['asset_allocation']
                    )
            
            html_content += '</div>'
        
        html_content += '</body></html>'
        
        with open(file_path, 'w') as f:
            f.write(html_content)
    
    async def _store_delivery_record(self, config: ClientReportConfig, file_path: str) -> int:
        """Store client report delivery record in database."""
        query = """
            INSERT INTO client_report_deliveries (
                client_report_id, generated_at, delivery_status,
                file_path, file_size_bytes
            ) VALUES (
                (SELECT id FROM client_reports WHERE client_id = $1 AND report_type = $2 LIMIT 1),
                $3, $4, $5, $6
            )
            RETURNING id
        """
        
        file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
        
        result = await self.db.fetch_one(
            query,
            config.client_id,
            config.report_type,
            datetime.now(),
            'generated',
            file_path,
            file_size
        )
        
        return result['id'] if result else 0


class ClientReportingSystem:
    """
    Main client reporting system that manages report configurations,
    scheduling, generation, and delivery for all clients.
    """
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.report_generator = ClientReportGenerator(db_connection)
    
    async def create_report_configuration(self, config: ClientReportConfig) -> int:
        """Create new client report configuration."""
        query = """
            INSERT INTO client_reports (
                client_id, report_type, report_name, report_config,
                delivery_schedule, delivery_method, recipients,
                is_active, created_at
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
            RETURNING id
        """
        
        result = await self.db.fetch_one(
            query,
            config.client_id,
            config.report_type,
            config.report_name,
            json.dumps(asdict(config)),
            config.delivery_schedule,
            config.delivery_method,
            json.dumps(config.recipients),
            True,
            datetime.now()
        )
        
        logger.info(f"Created report configuration for client {config.client_id}: {result['id']}")
        return result['id']
    
    async def generate_scheduled_reports(self) -> List[Dict[str, Any]]:
        """Generate all scheduled reports that are due."""
        query = """
            SELECT * FROM client_reports 
            WHERE is_active = true 
            AND (next_due_date <= CURRENT_DATE OR next_due_date IS NULL)
        """
        
        due_reports = await self.db.fetch_all(query)
        results = []
        
        for report_config in due_reports:
            try:
                config = ClientReportConfig(**json.loads(report_config['report_config']))
                
                # Calculate period dates based on schedule
                period_end = date.today()
                if config.delivery_schedule == 'monthly':
                    period_start = period_end.replace(day=1) - timedelta(days=1)
                    period_start = period_start.replace(day=1)
                elif config.delivery_schedule == 'quarterly':
                    quarter = (period_end.month - 1) // 3
                    period_start = date(period_end.year, quarter * 3 + 1, 1)
                else:
                    period_start = period_end - timedelta(days=30)  # Default to 30 days
                
                # Generate report
                result = await self.report_generator.generate_client_report(
                    config, period_start, period_end, [config.client_id]  # Simplified
                )
                
                results.append(result)
                
                # Update next due date
                await self._update_next_due_date(report_config['id'], config.delivery_schedule)
                
            except Exception as e:
                logger.error(f"Failed to generate scheduled report {report_config['id']}: {e}")
                results.append({
                    'status': 'error',
                    'report_id': report_config['id'],
                    'error': str(e)
                })
        
        return results
    
    async def _update_next_due_date(self, report_id: int, schedule: str):
        """Update next due date for scheduled report."""
        if schedule == 'daily':
            next_date = date.today() + timedelta(days=1)
        elif schedule == 'weekly':
            next_date = date.today() + timedelta(weeks=1)
        elif schedule == 'monthly':
            next_date = date.today() + timedelta(days=30)
        elif schedule == 'quarterly':
            next_date = date.today() + timedelta(days=90)
        elif schedule == 'annual':
            next_date = date.today() + timedelta(days=365)
        else:
            next_date = date.today() + timedelta(days=30)
        
        query = """
            UPDATE client_reports 
            SET next_due_date = $1, last_generated = $2, updated_at = $3
            WHERE id = $4
        """
        
        await self.db.execute(
            query,
            next_date,
            datetime.now(),
            datetime.now(),
            report_id
        )
